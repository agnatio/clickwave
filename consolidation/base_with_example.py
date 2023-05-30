import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import win32gui
import os
import pyautogui as pag
from datetime import datetime
# from jinja2 import Environment, FileSystemLoader

class BaseForm:
    def __init__(self):
        # set the initial position of the form
        x_pos = 700
        y_pos = 180
        self.root = tk.Tk()
        self.root.title("Abstract")
        self.root.wm_attributes('-topmost', True) # set the form to always be on top
        self.WIDTH = 15
        self.root.geometry(f"+{x_pos}+{y_pos}") # set the initial position of the form
        # self.root.resizable(False, False) # disable resizing the form
        self.root.iconbitmap('invoice_rule/fox.ico')
        self.folder_path = os.path.dirname(os.path.abspath(__file__)) # get the path of the folder where the script is located

        self.labels = {}
        self.entries = {}
        self.form_data = {}
        self.i = 0
        self.updated_status = False
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        # self.description = None
        self.source_file = 'Empty'
        self.create_start_widgets()

    def create_start_widgets(self):
        # add a button to open the file with filedialog
        self.open_file_button = ttk.Button(self.root, text="Open File", command=self.process_file)
        self.open_file_button.grid(row=0, column=0, padx= 5, pady = 5, sticky='e')
        self.create_file_name_label()

    def create_file_name_label(self):
        # create the label with the file name
        file_label = os.path.basename(self.source_file)
        self.file_name_label = ttk.Label(self.root, text=file_label, width=self.WIDTH)
        self.file_name_label.grid(row=0, column=1, padx= 5, pady = 5, sticky='w')

    def get_file_path(self):
        print(self.source_file)
        return filedialog.askopenfilename(initialdir=f"{self.folder_path}", title="Select a File", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
        

    def process_file(self):
        self.source_file = self.get_file_path() # get the file path
        self.data = self.load_data() # load the data from the file
        self.remove_widgets() # Remove previously placed buttons
        self.create_entries(self.data) # create the widgets
        self.create_base_butons() # create buttons to submit the form, exit the application, and load the form
        self.create_checkbox_log() 
        self.create_file_name_label() # create the label with the file name
        self.place_entries(self.data) # place the widgets
        self.place_stat_info() # place the label with the current row
        self.place_base_buttons() # place the buttons
        self.place_checkbox()
        self.additional_widgets()
        self.place_additional_widgets()
        self.update_external()

    def load_data(self) -> None:
        # load the data from the file
        data_list = []
        workbook = openpyxl.load_workbook(filename=self.source_file) # load the file
        worksheet = workbook.active # get the active worksheet
        headers = [cell.value for cell in next(worksheet.iter_rows())]  # read column headers
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # create dictionary with column names as keys
            data_list.append(dict(zip(headers, row)))
        # add a header 'updated' with value False to each dictionary
        print(headers)
        if 'updated' not in headers:
            for row in data_list:
                row['updated'] = False
        else:
            self.updated_status = True
            print("Updated status is True")
        self.current_row = 0
        print(data_list[self.current_row])
        # if the number of columns is greater than 20, don't load the form and put a message
        if len(data_list[0].keys()) > 20:
            messagebox.showerror("Error", "The number of columns is greater than 20")
            return
        return data_list

        
    def create_entries(self, data):
        # create the widgets
        for key, value in data[0].items():
            label = ttk.Label(self.root, text=key)
            entry = ttk.Entry(self.root, width=self.WIDTH)
            entry.insert(0, value)
            self.labels[key] = label
            self.entries[key] = entry

    def create_base_butons(self):
        # create buttons to submit the form, exit the application, and load the form
        self.submit_button = ttk.Button(self.root, text="Submit", command=self.submit_form)
        self.exit_button = ttk.Button(self.root, text="Exit", command=self.on_closing)
        self.forw_button = ttk.Button(self.root, text=">", command=self.next_row)
        self.back_button = ttk.Button(self.root, text="<", command=self.prev_row)

    def create_checkbox_log(self):
        self.checkbox_log_var = tk.IntVar()
        self.checkbox_log = ttk.Checkbutton(self.root, text="Keep log", variable=self.checkbox_log_var)
        self.checkbox_all_var = tk.IntVar()
        self.checkbox_all = ttk.Checkbutton(self.root, text="Process all unprocessed", variable=self.checkbox_all_var)


    def update_stat_info(self):
        self.window_info_label.config(text=f"{self.current_row+1} of {len(self.data)}")

    def remove_widgets(self):
        # Remove previously placed buttons
        for widget in self.root.winfo_children():
            if widget != self.open_file_button:
                widget.destroy()

    def place_entries(self, data):
        # place the entry widgets
        for self.i, key in enumerate(data[0].keys()):
            self.labels[key].grid(row=self.i+1, column=0, padx=5, pady=5, sticky='e')
            self.entries[key].grid(row=self.i+1, column=1, columnspan=5, padx=5, pady=5, sticky='we')


    def place_stat_info(self):    
        # add a lable to show the information about the window n's out of total number of rows
        self.window_info_label = ttk.Label(self.root, text=f"{self.current_row+1} of {len(self.data)}")
        self.window_info_label.grid(row=self.i+5, column=1, padx= 5, pady = 5, sticky='w')

    def place_base_buttons(self):
        # place the buttons
        self.exit_button.grid(row=self.i+6, column=0, padx = 5, pady = 5, sticky='e')
        self.submit_button.grid(row=self.i+6, column=1, padx = 5, pady = 5, sticky='w')
        self.back_button.grid(row=self.i+6, column=4, padx = 5, pady = 5, sticky='w')
        self.forw_button.grid(row=self.i+6, column=5,padx = 5, pady = 5, sticky='w')

    def place_checkbox(self):
        self.checkbox_log.grid(row=self.i+6, column=2, padx=5, pady=5, sticky='we')
        self.checkbox_all.grid(row=self.i+6, column=3, padx=5, pady=5, sticky='we')


        
    def populate_form(self):
        # populate the form with the data from the current row
        for el in self.data[self.current_row]:
            self.entries[el].delete(0, tk.END)
            self.entries[el].insert(0, self.data[self.current_row][el])
    
    
    def next_row(self):
        if self.current_row < len(self.data) - 1:
            self.current_row += 1
        # else begin at the first row
        else:
            self.current_row = 0
        self.update_stat_info()
        self.populate_form()
        self.update_external()
        
    def prev_row(self):
        if self.current_row > 0:
            self.current_row -= 1
        # else begin at the last row
        else:
            self.current_row = len(self.data) - 1
        self.update_stat_info()
        self.populate_form()
        self.update_external()

    def whatiswindow_function(self):
        # Temporarily hide the Tkinter window
        self.root.withdraw()
        # Get the handle of the active window
        hwnd = win32gui.GetForegroundWindow()
        # Get the title of the active window
        self.window_title = win32gui.GetWindowText(hwnd)
        print(self.window_title)

        # Restore the Tkinter window
        self.root.deiconify()
        return self.window_title
    
    def activate_window_with_name(self, name):
        def callback(hwnd, hwnds):
            if win32gui.IsWindowVisible(hwnd) and name.lower() in win32gui.GetWindowText(hwnd).lower():
                win32gui.SetForegroundWindow(hwnd)
                return

        win32gui.EnumWindows(callback, None)
    
    def activate_foreground_window(self):
        # Get the handle of the foreground window
        foreground_window = win32gui.GetForegroundWindow()

        # Activate the foreground window
        win32gui.SetForegroundWindow(foreground_window)

    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            # update the file with 'updated' field
            if self.updated_status:
                # create a new file with the same name as the source file
                new_file_name = self.source_file + "_updated.xlsx"
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                # write the headers
                for i, key in enumerate(self.data[0].keys()):
                    worksheet.cell(row=1, column=i+1).value = key
                # write the data
                for i, row in enumerate(self.data):
                    for j, key in enumerate(row.keys()):
                        worksheet.cell(row=i+2, column=j+1).value = row[key]
                workbook.save(new_file_name)

            self.root.destroy()

    def submit_status_true(self):
        self.updated_status = True
    
 
    def submit_form(self):
        ...

    def renew_and_place(self):
        ...

    def update_external(self):
        ...
    
    def mainloop(self):
        self.root.mainloop()
    

class ConsolidationPost(BaseForm):
    def __init__(self, system_instance):
        super().__init__()
        self.root.title("Consolidation")
        self.sleep = 0.5
        self.system_instance = system_instance

    def additional_widgets(self):
        self.horisontal_line = ttk.Separator(self.root, orient='horizontal')
        self.description = ttk.Label(self.root, text="Description")
        self.description_entry = ttk.Entry(self.root, width=self.WIDTH)
        self.labels['Description'] = self.description
        self.entries['Description'] = self.description_entry
        

    def place_additional_widgets(self):
        # place horisontal line
        self.horisontal_line.grid(row=self.i+7, column=0, columnspan=5, padx=5, pady=5, sticky='we')
        # place the entry widgets
        self.description.grid(row=self.i+8, column=0, padx=5, pady=5, sticky='e')
        self.description_entry.grid(row=self.i+8, column=1, columnspan=5, padx=5, pady=5, sticky='we')

    def update_external(self):
        # update description field with concatinated values
        self.description_entry.delete(0, tk.END)
        # make a string from self.data not longer than 20 symbols
        company = str(self.data[self.current_row]['Mapping'])
        period_type = str(self.data[self.current_row]['Amount_Type'])
        period = str(self.data[self.current_row]['Period'])
        insert_desc = f"Enter data for {company}, for {period_type}, for {period}"
        self.description_entry.insert(0, insert_desc)

    def submit_form(self):
        print("Additional submit function")
        print(f"{self.checkbox_log_var.get()} is the value of the checkbox")
        # create pyautogui question box 'Are you sure you want to submit?' with yes and no option
        submit_answer = messagebox.askyesno("Submit", "Are you sure you want to submit?")
        if submit_answer:
            print("Submitting")
            # move mouse to coordinates 100, 100 and click left mouse button
            pag.moveTo(230, 145, duration=self.sleep*2)
            pag.sleep(self.sleep)
            pag.click()
            # type self.data[self.current_row]['Mapping'] and press tab
            pag.sleep(self.sleep)
            pag.typewrite(str(self.data[self.current_row]['Mapping']))
            pag.sleep(self.sleep)
            pag.press('tab')
            pag.press('tab')
            pag.typewrite(str(self.data[self.current_row]['Amount_Type']))
            pag.press('tab')
            pag.typewrite(str(self.data[self.current_row]['Period']))
            pag.press('tab')
            pag.typewrite(str(self.data[self.current_row]['Standard_Period']))
            submit_answer = messagebox.askyesno("Submit", "Do you like filling?")
            if submit_answer:
                self.activate_foreground_window()
                pag.sleep(self.sleep)
                self.activate_window_with_name(self.system_instance)
                pag.press('enter')
                pag.sleep(self.sleep)
                # ipdate 'updated' field in self.data with current time with seconds
                self.data[self.current_row]['updated'] = datetime.now().strftime("%Y.%m.%d %H:%M:%S")
                # refresh the 'updated' field in the form
                self.populate_form()
                # update the form with the next row
                self.next_row()
                self.submit_status_true()
            else:
                pass
        

        


if __name__ == '__main__':
    

    app = ConsolidationPost("Diagnostic")
    app.mainloop()